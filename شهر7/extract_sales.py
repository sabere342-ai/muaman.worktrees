import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\saber\OneDrive\Desktop\ادارة_محل_مؤمن\شهر7\شيت_ادارة_محل_مؤمن_مطور_حديث_شهر7.xlsx', data_only=True)
ws = wb['المبيعات']

count = 0
for r in range(2, ws.max_row + 1):
    date_val = ws.cell(r, 2).value
    product = ws.cell(r, 3).value
    barcode = ws.cell(r, 4).value
    qty = ws.cell(r, 5).value
    sale_price = ws.cell(r, 6).value
    cost_price = ws.cell(r, 8).value

    if date_val is None or product is None:
        continue

    if isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        continue

    if isinstance(barcode, (int, float)):
        barcode = str(int(barcode))

    count += 1
    print(f"['{date_str}', '{product}', '{barcode}', {qty}, {sale_price}, {cost_price}],")

print(f"\n// Total rows: {count}")
